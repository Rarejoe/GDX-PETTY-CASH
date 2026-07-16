"""
Petty Cash Request System (Postgres / Supabase version)
---------------------------------------------------------
A small Flask web app that replaces paper petty cash request forms.

- Anyone with the link can submit a request (multiple expense lines + drawn signature).
- Data is stored permanently in a Postgres database (Supabase free tier).
- Approver gets an email notification when a new request is submitted.
- A dashboard lists all requests and lets an approver mark them Approved / Rejected / Paid.
- Approver actions are protected by a single shared password (set via APPROVER_PASSWORD).

Requires the DATABASE_URL environment variable (your Supabase connection string).

Run locally:
    pip install -r requirements.txt
    python app.py
Then open http://127.0.0.1:5000 in your browser.
"""
from supabase import create_client
import uuid
import os
import datetime
from zoneinfo import ZoneInfo
from functools import wraps

from io import BytesIO
from flask import send_file
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle,
    Paragraph, Spacer
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

import psycopg
from psycopg.rows import dict_row
import requests
from flask import (
    Flask, render_template, request, redirect,
    url_for, session, flash, g
)

DATABASE_URL = os.environ.get("DATABASE_URL")

app = Flask(__name__)
import os

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

supabase = create_client(
    SUPABASE_URL,
    SUPABASE_KEY
)
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-change-me")
APPROVER_PASSWORD = os.environ.get("APPROVER_PASSWORD", "changeme123")
RESEND_API_KEY = os.environ.get("RESEND_API_KEY")
APPROVER_EMAIL = os.environ.get("APPROVER_EMAIL")


def send_approver_notification(ref_no, requester, gross_total):
    """Email the approver when a new request comes in. Fails silently if
    email isn't configured or the send fails, so it never blocks a submission."""
    if not RESEND_API_KEY or not APPROVER_EMAIL:
        print("EMAIL SKIPPED: RESEND_API_KEY or APPROVER_EMAIL not set")
        return
    try:
        resp = requests.post(
            "https://api.resend.com/emails",
            headers={"Authorization": f"Bearer {RESEND_API_KEY}"},
            json={
                "from": "GDX Petty Cash <onboarding@resend.dev>",
                "to": [e.strip() for e in APPROVER_EMAIL.split(",") if e.strip()],
                "subject": f"New petty cash request {ref_no}",
                "html": (
                    f"<p><strong>{requester}</strong> submitted a new petty cash request.</p>"
                    f"<p>Reference: <strong>{ref_no}</strong><br>"
                    f"Amount: <strong>₦{gross_total:,.2f}</strong></p>"
                    f"<p><a href=\"https://gdx-petty-cash.onrender.com/login\">Log in to review it</a></p>"
                ),
            },
            timeout=10,
        )
        print(f"EMAIL RESPONSE: status={resp.status_code} body={resp.text}")
    except Exception as ex:
        print(f"EMAIL FAILED: {ex}")


# ---------------------------------------------------------------------------
# Database helpers
# ---------------------------------------------------------------------------

def get_db():
    if "db" not in g:
        g.db = psycopg.connect(
            DATABASE_URL,
            row_factory=dict_row
        )
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db():
    db = psycopg.connect(DATABASE_URL)
    cur = db.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS requests (
            id SERIAL PRIMARY KEY,
            ref_no TEXT UNIQUE NOT NULL,
            request_date TEXT NOT NULL,
            requester TEXT NOT NULL,
            department TEXT,
            purpose TEXT,
            signature_name TEXT NOT NULL,
            signature_image TEXT,
            receipt_url TEXT,
            signed_on TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'Pending',
            approver_name TEXT,
            approved_on TEXT,
            gross_total REAL NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL
        );
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS line_items (
            id SERIAL PRIMARY KEY,
            request_id INTEGER NOT NULL REFERENCES requests(id),
            line_date TEXT,
            description TEXT NOT NULL,
            amount REAL NOT NULL
        );
    """)
    db.commit()
    cur.close()
    db.close()


def next_ref_no(db):
    cur = db.cursor()
    cur.execute("""
        SELECT ref_no FROM requests
        WHERE ref_no LIKE 'PCR-%%'
        ORDER BY id DESC LIMIT 1
    """)
    row = cur.fetchone()
    cur.close()
    if row is None:
        return "PCR-0001"
    last_num = int(row["ref_no"].split("-")[1])
    return f"PCR-{last_num + 1:04d}"


# ---------------------------------------------------------------------------
# Auth helper for approver-only pages
# ---------------------------------------------------------------------------

def approver_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not session.get("is_approver"):
            flash("Please log in as an approver to do that.", "error")
            return redirect(url_for("login", next=request.path))
        return view_func(*args, **kwargs)
    return wrapped


# ---------------------------------------------------------------------------
# Public routes
# ---------------------------------------------------------------------------

@app.route("/", methods=["GET"])
def new_request_form():
    db = get_db()
    ref_no = next_ref_no(db)
    today = datetime.date.today().isoformat()
    return render_template("request_form.html", ref_no=ref_no, today=today)
    
@app.route("/archive/<int:request_id>", methods=["POST"])
@approver_required
def archive_request(request_id):
    db = get_db()
    cur = db.cursor()

    cur.execute(
        "UPDATE requests SET archived = TRUE WHERE id = %s",
        (request_id,)
    )

    db.commit()
    cur.close()

    flash("Request archived.", "success")
    return redirect(url_for("dashboard"))
    
@app.route("/download/<int:request_id>")
@approver_required
def download_request(request_id):
    db = get_db()
    cur = db.cursor()

    cur.execute("SELECT * FROM requests WHERE id = %s", (request_id,))
    req = cur.fetchone()

    cur.execute("""
        SELECT *
        FROM line_items
        WHERE request_id = %s
        ORDER BY id
    """, (request_id,))
    items = cur.fetchall()

    cur.close()

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>GDX EQUIP</b>", styles["Title"]))
    elements.append(Paragraph("<b>PETTY CASH REQUEST</b>", styles["Heading2"]))
    elements.append(Spacer(1, 0.2 * inch))

    elements.append(Paragraph(f"<b>Reference:</b> {req['ref_no']}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Status:</b> {req['status']}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Requester:</b> {req['requester']}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Request Date:</b> {req['request_date']}", styles["Normal"]))
    elements.append(Spacer(1, 0.2 * inch))

    data = [["Date", "Description", "Amount"]]

    for item in items:
        data.append([
            str(item["line_date"] or ""),
            item["description"],
            f"₦{item['amount']:,.2f}"
        ])

    data.append([
        "",
        "GROSS TOTAL",
        f"₦{req['gross_total']:,.2f}"
    ])

    table = Table(data, colWidths=[100, 260, 100])

    table.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.5, colors.black),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("BACKGROUND", (0,-1), (-1,-1), colors.beige),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("ALIGN", (2,0), (2,-1), "RIGHT"),
        ("BOTTOMPADDING", (0,0), (-1,0), 8),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 0.3 * inch))

    elements.append(Paragraph(
        f"<b>Requester Signature:</b> {req['signature_name']}",
        styles["Normal"]
    ))

    if req["approver_name"]:
        elements.append(Paragraph(
            f"<b>Approved By:</b> {req['approver_name']}",
            styles["Normal"]
        ))

    doc.build(elements)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{req['ref_no']}.pdf",
        mimetype="application/pdf"
    )
@app.route("/archived")
@approver_required
def archived():
    db = get_db()
    cur = db.cursor()

    cur.execute("""
        SELECT *
        FROM requests
        WHERE archived = TRUE
        ORDER BY created_at DESC
    """)

    rows = cur.fetchall()
    cur.close()

    return render_template("archived.html", requests=rows)
    
@app.route("/restore/<int:request_id>", methods=["POST"])
@approver_required
def restore_request(request_id):
    db = get_db()
    cur = db.cursor()

    cur.execute(
        "UPDATE requests SET archived = FALSE WHERE id = %s",
        (request_id,)
    )

    db.commit()
    cur.close()

    flash("Request restored.", "success")
    return redirect(url_for("archived"))    
    
@app.route("/submit", methods=["POST"])
def submit_request():
    db = get_db()
    cur = db.cursor()

    requester = request.form.get("requester", "").strip()
    request_date = request.form.get("request_date", "").strip()
    signature_name = request.form.get("signature_name", "").strip()
    signature_image = request.form.get("signature_image", "").strip()

    receipt = request.files.get("receipt")
    receipt_url = None

    line_dates = request.form.getlist("line_date[]")
    line_descs = request.form.getlist("line_desc[]")
    line_amounts = request.form.getlist("line_amount[]")

    errors = []

    if not requester:
        errors.append("Requester's name is required.")

    if not signature_name:
        errors.append("Your name is required to confirm the request.")

    if not signature_image:
        errors.append("Please sign the request before submitting.")

    line_items = []

    for d, desc, amt in zip(line_dates, line_descs, line_amounts):
        desc = desc.strip()

        if not desc and not amt.strip():
            continue

        try:
            amount = float(amt)
        except ValueError:
            amount = 0.0

        if desc and amount > 0:
            line_items.append({
                "date": d.strip(),
                "desc": desc,
                "amount": amount
            })

    if not line_items:
        errors.append("Please add at least one valid expense line (description + amount).")

    if errors:
        for e in errors:
            flash(e, "error")

        ref_no = next_ref_no(db)
        today = datetime.date.today().isoformat()

        return render_template(
            "request_form.html",
            ref_no=ref_no,
            today=today,
            form=request.form
        )

    gross_total = sum(i["amount"] for i in line_items)
    ref_no = next_ref_no(db)
    now = datetime.datetime.now(ZoneInfo("Africa/Lagos"))
    signed_on = now.strftime("%d %b %Y %I:%M %p")
    created_at = now.isoformat()

    if receipt and receipt.filename:
        file_ext = receipt.filename.rsplit(".", 1)[-1].lower()
        filename = f"{uuid.uuid4()}.{file_ext}"
        file_path = f"{ref_no}/{filename}"

       supabase.storage.from_("receipts").upload(
            file_path,
            receipt.read(),
            {"content-type": receipt.content_type}
        )

        receipt_url = file_path
        print("Receipt uploaded.URL:", receipt_url)

    cur.execute("""
        INSERT INTO requests
            (ref_no, request_date, requester, department, purpose,
             signature_name, signature_image, receipt_url, signed_on,
             status, gross_total, created_at)
        VALUES (%s, %s, %s, NULL, NULL, %s, %s, %s, %s,
                'Pending', %s, %s)
        RETURNING id
    """, (
        ref_no,
        request_date,
        requester,
        signature_name,
        signature_image,
        receipt_url,
        signed_on,
        gross_total,
        created_at
    ))

    request_id = cur.fetchone()["id"]

    for item in line_items:
        cur.execute("""
            INSERT INTO line_items
            (request_id, line_date, description, amount)
            VALUES (%s, %s, %s, %s)
        """, (
            request_id,
            item["date"],
            item["desc"],
            item["amount"]
        ))

    db.commit()
    cur.close()

    send_approver_notification(ref_no, requester, gross_total)

    return redirect(url_for("confirmation", ref_no=ref_no))


@app.route("/confirmation/<ref_no>")
def confirmation(ref_no):
    return render_template("confirmation.html", ref_no=ref_no)


# ---------------------------------------------------------------------------
# Approver login
# ---------------------------------------------------------------------------

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        password = request.form.get("password", "")
        if password == APPROVER_PASSWORD:
            session["is_approver"] = True
            flash("Logged in as approver.", "success")
            next_url = request.args.get("next") or url_for("dashboard")
            return redirect(next_url)
        flash("Incorrect password.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("is_approver", None)
    flash("Logged out.", "success")
    return redirect(url_for("new_request_form"))


# ---------------------------------------------------------------------------
# Dashboard (approver view)
# ---------------------------------------------------------------------------

@app.route("/dashboard")
@approver_required
def dashboard():
    db = get_db()
    cur = db.cursor()

    status_filter = request.args.get("status", "All")

    if status_filter != "All":
        cur.execute("""
            SELECT *
            FROM requests
            WHERE status = %s
              AND archived = FALSE
            ORDER BY created_at DESC
        """, (status_filter,))
    else:
        cur.execute("""
            SELECT *
            FROM requests
            WHERE archived = FALSE
            ORDER BY created_at DESC
        """)

    rows = cur.fetchall()

    cur.execute("""
        SELECT
            status,
            COUNT(*) AS count,
            COALESCE(SUM(gross_total), 0) AS total
        FROM requests
        WHERE archived = FALSE
        GROUP BY status
    """)
    totals = cur.fetchall()

    cur.close()

    return render_template(
        "dashboard.html",
        requests=rows,
        totals=totals,
        status_filter=status_filter
    )


@app.route("/request/<int:request_id>")
@approver_required
def request_detail(request_id):
    db = get_db()
    cur = db.cursor()
    cur.execute("SELECT * FROM requests WHERE id = %s", (request_id,))
    req = cur.fetchone()
    if req is None:
        flash("Request not found.", "error")
        cur.close()
        return redirect(url_for("dashboard"))
    cur.execute(
        "SELECT * FROM line_items WHERE request_id = %s ORDER BY id",
        (request_id,)
    )
    items = cur.fetchall()
    cur.close()
    return render_template("detail.html", req=req, items=items)


@app.route("/request/<int:request_id>/update_status", methods=["POST"])
@approver_required
def update_status(request_id):
    db = get_db()
    cur = db.cursor()

    new_status = request.form.get("status")
    approver_name = request.form.get("approver_name", "").strip()

    if new_status not in ("Pending", "Approved", "Rejected", "Paid"):
        flash("Invalid status.", "error")
        cur.close()
        return redirect(url_for("request_detail", request_id=request_id))

    approved_on = None
    paid_on = None

    if new_status in ("Approved", "Rejected"):
        approved_on = datetime.datetime.now().strftime("%d %b %Y %I:%M %p")

    elif new_status == "Paid":
        paid_on = datetime.datetime.now().strftime("%d %b %Y %I:%M %p")

    cur.execute("""
        UPDATE requests
        SET
            status = %s,
            approver_name = %s,
            approved_on = COALESCE(%s, approved_on),
            paid_on = COALESCE(%s, paid_on)
        WHERE id = %s
    """, (
        new_status,
        approver_name,
        approved_on,
        paid_on,
        request_id
    ))

    db.commit()
    cur.close()

    flash(f"Request marked as {new_status}.", "success")
    return redirect(url_for("request_detail", request_id=request_id))

# ---------------------------------------------------------------------------
# Excel export (for auditing)
# ---------------------------------------------------------------------------

@app.route("/export")
@approver_required
def export_excel():
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    db = get_db()
    cur = db.cursor()

    status_filter = request.args.get("status", "All")
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()

    query = """
        SELECT
            r.ref_no, r.requester, r.request_date,
            li.line_date, li.description, li.amount,
            r.gross_total, r.status, r.signed_on,
            r.approved_on, r.paid_on, r.approver_name
        FROM requests r
        JOIN line_items li ON li.request_id = r.id
        WHERE r.archived = FALSE
    """
    params = []

    if status_filter and status_filter != "All":
        query += " AND r.status = %s"
        params.append(status_filter)
    if start_date:
        query += " AND r.request_date >= %s"
        params.append(start_date)
    if end_date:
        query += " AND r.request_date <= %s"
        params.append(end_date)

    query += " ORDER BY r.created_at DESC, li.id"

    cur.execute(query, tuple(params))
    rows = cur.fetchall()
    cur.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Petty Cash Requests"

    headers = [
        "Reference No.", "Requester", "Request Date",
        "Expense Date", "Description", "Amount",
        "Gross Total", "Status", "Submitted On",
        "Approved On", "Paid On", "Approver Name",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="12294D", end_color="12294D", fill_type="solid")
    for col_num, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r["ref_no"], r["requester"], r["request_date"],
            r["line_date"], r["description"], r["amount"],
            r["gross_total"], r["status"], r["signed_on"],
            r["approved_on"], r["paid_on"], r["approver_name"],
        ])

    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename_bits = ["petty_cash_export"]
    if status_filter != "All":
        filename_bits.append(status_filter.lower())
    if start_date:
        filename_bits.append(start_date)
    if end_date:
        filename_bits.append(end_date)
    filename = "_".join(filename_bits) + ".xlsx"

    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

init_db()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
